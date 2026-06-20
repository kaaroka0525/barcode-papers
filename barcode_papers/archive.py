"""마음에 든 논문을 로컬에 정리: PDF 저장 + 엑셀 리스트업.

직전 다이제스트(data/last_digest.json) 기준으로 번호 또는 DOI를 지정합니다.

사용 예:
    python -m barcode_papers.archive 1 3        # 직전 메일의 1, 3번 논문
    python -m barcode_papers.archive all        # 직전 메일 전부
    python -m barcode_papers.archive 10.1016/...  # DOI 직접 지정

생성물:
    data/archive/<...>.pdf       무료(OA) PDF (가능한 경우)
    data/archive.xlsx            누적 리스트 (DOI 기준 중복 제거)

이후 Google Drive 업로드/시트화는 별도 단계에서 수행합니다.
"""
import argparse
import datetime as dt
import json
import logging
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

from . import pdfs
from .config import DATA_DIR, load_config
from .models import Paper
from .run import LAST_DIGEST_PATH

log = logging.getLogger("barcode_papers.archive")

ARCHIVE_DIR = DATA_DIR / "archive"
ARCHIVE_XLSX = DATA_DIR / "archive.xlsx"

HEADERS = [
    "추가날짜", "제목", "저자", "저널", "IF", "IF출처", "연도",
    "DOI", "원문링크", "한양대프록시", "PDF파일", "키워드", "요약",
]


def _load_last_digest():
    if not LAST_DIGEST_PATH.exists():
        log.error("직전 다이제스트가 없습니다(%s). 먼저 run을 실행하세요.", LAST_DIGEST_PATH)
        return []
    with open(LAST_DIGEST_PATH, "r", encoding="utf-8") as f:
        return [Paper.from_dict(d) for d in json.load(f)]


def _select(papers, tokens):
    """tokens: ['all'] 또는 번호들 또는 DOI들."""
    if not tokens or tokens == ["all"]:
        return list(papers)
    chosen = []
    for t in tokens:
        if t.isdigit():
            i = int(t) - 1
            if 0 <= i < len(papers):
                chosen.append(papers[i])
            else:
                log.warning("번호 %s 는 범위를 벗어났습니다(1~%d).", t, len(papers))
        else:
            match = next((p for p in papers if (p.doi or "").lower() == t.lower()), None)
            if match:
                chosen.append(match)
            else:
                log.warning("DOI '%s' 를 직전 다이제스트에서 찾지 못했습니다.", t)
    return chosen


def _existing_keys():
    if not ARCHIVE_XLSX.exists():
        return set()
    wb = load_workbook(ARCHIVE_XLSX)
    ws = wb.active
    keys = set()
    doi_col = HEADERS.index("DOI")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > doi_col and row[doi_col]:
            keys.add(str(row[doi_col]).lower())
    return keys


def _append_rows(rows):
    if ARCHIVE_XLSX.exists():
        wb = load_workbook(ARCHIVE_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "논문리스트"
        ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    wb.save(ARCHIVE_XLSX)


def _rclone_path():
    return shutil.which("rclone") or str(Path.home() / ".local/bin/rclone")


def upload_to_drive(config, pdf_names):
    """rclone로 엑셀 + (해당) PDF를 구글 드라이브 폴더에 업로드."""
    rclone = _rclone_path()
    if not Path(rclone).exists() and not shutil.which("rclone"):
        log.warning("rclone 미설치 → 드라이브 업로드 생략. (로컬 정리는 완료됨)")
        return
    dest = f"{config.gdrive_remote}:{config.gdrive_folder}/"
    try:
        subprocess.run([rclone, "copy", str(ARCHIVE_XLSX), dest], check=True, timeout=120)
        log.info("드라이브 업로드: archive.xlsx → %s", dest)
        for name in pdf_names:
            p = ARCHIVE_DIR / name
            if p.exists():
                subprocess.run([rclone, "copy", str(p), dest + "pdf/"], check=True, timeout=180)
                log.info("드라이브 업로드: %s → %spdf/", name, dest)
    except subprocess.CalledProcessError as e:
        log.warning("rclone 업로드 실패: %s", e)
    except Exception as e:
        log.warning("드라이브 업로드 중 오류: %s", e)


def archive(tokens, do_upload=True):
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    config = load_config()
    papers = _load_last_digest()
    if not papers:
        return
    selected = _select(papers, tokens)
    if not selected:
        log.info("선택된 논문이 없습니다.")
        return

    existing = _existing_keys()
    today = dt.date.today().isoformat()
    rows = []
    new_pdf_names = []
    for p in selected:
        if p.key in existing:
            log.info("이미 아카이브됨, 건너뜀: %s", p.title[:50])
            continue

        pdf_name = ""
        path = pdfs.download_pdf(p, ARCHIVE_DIR)
        if path:
            pdf_name = Path(path).name
            new_pdf_names.append(pdf_name)
            log.info("PDF 저장: %s", pdf_name)
        else:
            log.info("PDF 없음(유료/비공개): %s", p.title[:50])

        doi_link = f"https://doi.org/{p.doi}" if p.doi else (p.landing_url or "")
        proxy = config.proxy_link(p.landing_url or doi_link)
        rows.append([
            today, p.title, ", ".join(p.authors[:8]), p.journal,
            p.impact_factor if p.impact_factor is not None else "",
            p.if_source, p.year or "", p.doi or "", doi_link, proxy,
            pdf_name, p.matched_keyword, p.summary,
        ])

    if rows:
        _append_rows(rows)
        log.info("엑셀에 %d편 추가 → %s", len(rows), ARCHIVE_XLSX)
    else:
        log.info("새로 추가된 논문이 없습니다.")

    if do_upload and (rows or new_pdf_names):
        upload_to_drive(config, new_pdf_names)


def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s", stream=sys.stdout)
    for name in ("sqlalchemy", "sqlalchemy.engine", "sqlalchemy.engine.Engine"):
        logging.getLogger(name).setLevel(logging.WARNING)
    parser = argparse.ArgumentParser(description="마음에 든 논문 로컬 아카이브")
    parser.add_argument("items", nargs="*", default=["all"],
                        help="번호(1 3) / DOI / all (기본: all)")
    parser.add_argument("--no-upload", action="store_true", help="드라이브 업로드 없이 로컬 정리만")
    args = parser.parse_args()
    archive(args.items or ["all"], do_upload=not args.no_upload)


if __name__ == "__main__":
    main()
